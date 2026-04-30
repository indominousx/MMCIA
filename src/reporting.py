from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_csv_outputs(output_dir: Path, outputs: dict[str, pd.DataFrame]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, frame in outputs.items():
        frame.to_csv(output_dir / f"{name}.csv", index=False)


def build_weekly_report(
    output_dir: Path,
    data_frames: dict[str, pd.DataFrame],
    analysis_date: pd.Timestamp,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "weekly_purchase_report.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        _executive_summary(data_frames, analysis_date).to_excel(
            writer, sheet_name="Executive Summary", index=False
        )
        data_frames.get("procurement_recommendations", pd.DataFrame()).to_excel(
            writer, sheet_name="Supplier PO Plan", index=False
        )
        _immediate_actions(data_frames).to_excel(writer, sheet_name="Immediate Actions", index=False)
        data_frames.get("stockout_alerts_21d", pd.DataFrame()).to_excel(
            writer, sheet_name="Stockout Alerts", index=False
        )
        data_frames.get("substitution_recommendations", pd.DataFrame()).to_excel(
            writer, sheet_name="Substitutions", index=False
        )
        _bom_trace(data_frames).to_excel(writer, sheet_name="BOM Demand Trace", index=False)
        data_frames.get("slow_moving_watchlist", pd.DataFrame()).to_excel(
            writer, sheet_name="Slow Moving Watchlist", index=False
        )
        data_frames.get("data_quality_issues", pd.DataFrame()).to_excel(
            writer, sheet_name="Data Quality Notes", index=False
        )
        data_frames.get("procurement_blocked_by_credit", pd.DataFrame()).to_excel(
            writer, sheet_name="Blocked By Credit", index=False
        )

        workbook = writer.book
        for sheet in workbook.worksheets:
            _format_sheet(sheet)
    return report_path


def _executive_summary(data_frames: dict[str, pd.DataFrame], analysis_date: pd.Timestamp) -> pd.DataFrame:
    credit = data_frames.get("credit_summary", pd.DataFrame())
    approved = data_frames.get("procurement_recommendations", pd.DataFrame())
    blocked = data_frames.get("procurement_blocked_by_credit", pd.DataFrame())
    alerts = data_frames.get("stockout_alerts_21d", pd.DataFrame())

    approved_value = approved["recommended_value_inr"].sum() if not approved.empty else 0.0
    blocked_value = blocked["recommended_value_inr"].sum() if not blocked.empty else 0.0
    remaining_credit = (
        credit["remaining_available_credit_inr"].iloc[0]
        if not credit.empty and "remaining_available_credit_inr" in credit
        else 0.0
    )
    credit_cap = credit["credit_cap_inr"].iloc[0] if not credit.empty and "credit_cap_inr" in credit else 0.0

    rows = [
        {"metric": "analysis_date", "value": analysis_date.date().isoformat()},
        {"metric": "approved_recommendation_lines", "value": len(approved)},
        {"metric": "approved_po_value_inr", "value": approved_value},
        {"metric": "blocked_by_credit_lines", "value": len(blocked)},
        {"metric": "blocked_by_credit_value_inr", "value": blocked_value},
        {"metric": "stockout_alert_lines", "value": len(alerts)},
        {"metric": "credit_cap_inr", "value": credit_cap},
        {"metric": "remaining_available_credit_inr", "value": remaining_credit},
    ]
    return pd.DataFrame(rows)


def _immediate_actions(data_frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    approved = data_frames.get("procurement_recommendations", pd.DataFrame())
    blocked = data_frames.get("procurement_blocked_by_credit", pd.DataFrame())
    frames = []
    if not approved.empty and "action_timing" in approved:
        frames.append(approved[approved["action_timing"] == "immediate"].copy())
    if not blocked.empty and "action_timing" in blocked:
        blocked_now = blocked[blocked["action_timing"] == "immediate"].copy()
        if not blocked_now.empty:
            frames.append(blocked_now)
    if not frames:
        return pd.DataFrame(columns=["action", "supplier_name", "material_id", "material_name", "credit_status"])
    immediate = pd.concat(frames, ignore_index=True, sort=False)
    immediate.insert(
        0,
        "action",
        immediate["credit_status"].apply(
            lambda status: "order now" if str(status).startswith("approved") else "escalate credit/substitution"
        ),
    )
    return immediate


def _bom_trace(data_frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    trace = data_frames.get("bom_exploded_order_demand", pd.DataFrame())
    if trace.empty:
        return trace
    columns = [
        "order_id",
        "client_id",
        "product_type",
        "box_size",
        "delivery_date",
        "material_id",
        "material_name",
        "bom_qty_per_box",
        "order_quantity",
        "raw_required_qty",
        "seasonal_multiplier",
        "seasonal_required_qty",
        "canonical_unit",
    ]
    columns = [column for column in columns if column in trace.columns]
    return trace[columns].head(250)


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 42))
